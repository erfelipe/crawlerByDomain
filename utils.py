# from sqlite3.dbapi2 import Cursor
from genericpath import exists
from sqlite3.dbapi2 import Error
from urllib.parse import urlparse
import crawDomain as craw 
import consts
import sqlite3
import unicodedata 
import os
from wordcloud import WordCloud 
from bs4 import BeautifulSoup
from urlextract import URLExtract 
from openpyxl import Workbook 
from datetime import datetime, timezone

keywords = []
caracteresAcentuados = ["á", "à", "ã", "é", "í", "ô", "ç"]
blackListDomains = []

def dropFile(arq):
    if (os.path.exists(arq)):
        print("Deleting file!... " + arq)
        os.remove(arq)

def createDataBase():
    conn = sqlite3.connect(consts.ARQ_DATABASE) 
    cursor = conn.cursor() 
    cursor.execute("""  CREATE TABLE if not exists crawlerByDomain(
                        id integer primary key autoincrement,
                        domain text, 
                        url text not null,
                        page text not null,
                        keys text not null,
                        quantkeys integer not null, 
                        titleFromPage text, 
                        keysFromPage text, 
                        langFromPage text, 
                        descFromPage text, 
                        authorFromPage text)
                    """) 
    conn.commit() 
    cursor.execute("""  CREATE TABLE if not exists counters (
                        id integer primary key autoincrement,
                        pagesVisited integer,
                        validDomains integer,
                        stamp datetime
                        )
                    """) 
    conn.commit() 
    conn.close() 

def insertDataInDB(url, page, keywordsFound): 
    keys = ', '.join(keywordsFound) 
    quantk =  len(keywordsFound) 
    domain = getDomainFromURL(url) 

    metadata = extractMetaFromPage(page)
    titleFromPage = metadata["title"]
    keysFromPage = metadata["keywords"] 
    langFromPage = metadata["lang"] 
    descFromPage = metadata["description"] 
    authorFromPage = metadata["author"] 

    conn = sqlite3.connect(consts.ARQ_DATABASE) 
    cursor = conn.cursor() 
    cursor.execute("""  INSERT INTO crawlerByDomain (domain, url, page, keys, quantkeys, titleFromPage, keysFromPage, langFromPage, descFromPage, authorFromPage) 
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ? ) """, (domain, url, page, keys, quantk, titleFromPage, keysFromPage, langFromPage, descFromPage, authorFromPage,))
    conn.commit()
    conn.close() 

def insertQuantsInDB(pagesVisited, validDomains):
    conn = sqlite3.connect(consts.ARQ_DATABASE) 
    cursor = conn.cursor() 
    dtime = datetime.now().strftime("%B %d, %Y %I:%M%p")
    cursor.execute("""  INSERT INTO counters(pagesVisited, validDomains, stamp) VALUES(?, ?, ?) """, (pagesVisited, validDomains, dtime,) )
    conn.commit() 
    conn.close() 

def cleanFinalFiles(): 
    newResultFiles = input("Clean result files for this process? (Y/N)").lower()
    if (newResultFiles == 'y'):
        if (exists("bd-result.xlsx")):
            os.remove("bd-result.xlsx")
        if (exists("keywordsFromVocabulary.png")):
            os.remove("keywordsFromVocabulary.png")
        if (exists("keys-for-cloud.txt")):
            os.remove("keys-for-cloud.txt")
        if (exists("quant_keys.xlsx")):
            os.remove("quant_keys.xlsx")
        if (exists("urls-to-visit.txt")):
            os.remove("urls-to-visit.txt")
        if (exists("urls-visited.txt")):
            os.remove("urls-visited.txt")

def repeteString(texto, n):
    """ Repete uma string n vezes e retorna sua concatenacao em nova string

    Args:
        texto (str): texto a ser repetido
        n (int): numero de vezes a ser repetida

    Returns:
        str: nova string com o numero de repeticao
    """    
    return ' '.join([texto for i in range(n)])

def fillKeywords():
    lista = open(consts.ARQ_KEYWORDS, 'r').read().splitlines()
    listaMinuscula = [x.lower().strip() for x in lista] 
    for l in listaMinuscula:
        keywords.append(l)

def fillBlackListDomains():
    lista = open(consts.ARQ_BLACKLISTDOMAIN, 'r').read().splitlines()
    listaMinuscula = [x.lower().strip() for x in lista] 
    for l in listaMinuscula:
        blackListDomains.append(l) 

def fillQueueWithSeeds(queueUrlsVisitar):
    seeds = loadFileLikeArray(consts.ARQ_SEEDS)
    for seed in seeds:
        queueUrlsVisitar.put(seed)

def initialize(queueUrlsVisitar, keywordsList, queueUrlsVisited):
    initBD()
    fillKeywords() 
    fillBlackListDomains()
    initQueueToVisit(queueUrlsVisitar)
    initkeywordsList(keywordsList)
    initQueueUrlsVisited(queueUrlsVisited)

def initBD():
    newDB = input("Create a new Dabatabase for this process? (y/n)").lower()
    if (newDB == 'y'):
        dropFile(consts.ARQ_DATABASE)
        createDataBase()
    newResultFiles = input("Create a new result files for analysis for this process? (Y/N)").lower()
    if (newResultFiles == 'y'):
        if (os.path.exists("quant_keys.xlsx")):
            os.remove("quant_keys.xlsx")
        if (os.path.exists("bd-result.xlsx")):
            os.remove("bd-result.xlsx")
        if (os.path.exists("keywordsFromVocabulary.png")):
            os.remove("keywordsFromVocabulary.png")
        if (os.path.exists("urls-to-visit.txt")):
            os.remove("urls-to-visit.txt")
        if (os.path.exists("urls-visited.txt")):
            os.remove("urls-visited.txt")
        if (os.path.exists("keys-for-cloud.txt")):
            os.remove("keys-for-cloud.txt")

def textNoAccent(texto):
    text = unicodedata.normalize('NFD', texto)\
           .encode('ascii', 'ignore')\
           .decode("utf-8")
    return text

def loadFileLikeArray(arq):
    """ Le o arquivo txt e gera um array em minusculas e sem espaços nas bordas
        Muito importante notar que caso tenha algum acento, o texto equivalente sem acento sera adicionado aa lista
    Returns:
        list -- Siglas tecnicas em minusculas 
    """    
    lista = open(arq, 'r').read().splitlines()
    listaMinuscula = [x.lower().strip() for x in lista] 
    for item in listaMinuscula:
        for letra in caracteresAcentuados:
            if letra in item:
                listaMinuscula.append(textNoAccent(item.lower()))
    return set(listaMinuscula)

def urlNotVisited(url, queueUrlsVisited):
    return url not in queueUrlsVisited

def keyswordsInDocument(data, keywordsList):
    keys = []
    for k in keywordsList:
        if k in data:
            keys.append(k)
    return keys

def isDataText(header): 
    try:
        cabecalho = header['content-type']
        if ('ASCII'.lower() in cabecalho.lower()) or ('ANSI'.lower() in cabecalho.lower()) or ('8859'.lower() in cabecalho.lower()) or ('UTF'.lower() in cabecalho.lower()) or ('text'.lower() in cabecalho.lower()) or ('html'.lower() in cabecalho.lower()) :
            return True
        else:
            return False
    except:
        return False

def urlWellFormat(url):
    try:
        result = urlparse(url)
        return all([result.scheme, result.netloc, result.path])
    except:
        return False

def loadUrlsVisited():
    urls = []
    urlFromDB = loadUrlsVisitedFromDB() 
    urlFromFile = loadUrlsVisitedFromFile() 
    for url in urlFromDB:
        urls.append(url)
    for url in urlFromFile:
        urls.append(url)    
    return urls

def loadUrlsVisitedFromDB():
    print("Loading urls from DB: " + consts.ARQ_DATABASE)
    try:
        conn = sqlite3.connect(consts.ARQ_DATABASE) 
    except Error:
        print(Error)
    cursor = conn.cursor() 
    cursor.execute("SELECT url FROM crawlerByDomain")
    result = cursor.fetchall()
    conn.close() 

    resp = []
    for url in result: 
        resp.append(url) 
    return resp

def loadUrlsVisitedFromFile():
    print("Loading urls visited: " + consts.ARQ_URLSVISITED)
    return loadArrayFromFile(consts.ARQ_URLSVISITED)

def getDomainFromURL(url):
    return urlparse(url).netloc
    
def isFileValid(url):
    return ( (url.rfind('.js') == -1) and (url.rfind('.jpg') == -1) and (url.rfind('.jpeg') == -1) and (url.rfind('.png') == -1) and (url.rfind('.css') == -1) and (url.rfind('.gif') == -1) and (url.rfind('.mp4') == -1) and (url.rfind('.ico') == -1) and (url.rfind('.svg') == -1) and (url.rfind('.pdf') == -1) and (url.rfind('.woff') == -1) and (url.rfind('.woff2') == -1) )

def saveUrlsVisited(urls):
    dropFile(consts.ARQ_URLSVISITED)
    print("Saving urls visited file: " + consts.ARQ_URLSVISITED)
    with open(consts.ARQ_URLSVISITED, "w") as f:
        for url in urls:
            f.write(str(url) +"\n")

def saveUrlsSearchers(urls):
    print("Saving urls searchers file: " + consts.ARQ_SEEDS_FROM_SEARCHERS)
    with open(consts.ARQ_SEEDS_FROM_SEARCHERS, "a") as f:
        for url in urls:
            f.write(str(url) +"\n")

def saveQueueToVisit(urls):
    dropFile(consts.ARQ_URLSTOVISIT)
    print("Saving queue to visit file: " + consts.ARQ_SEEDS_FROM_SEARCHERS)
    with open(consts.ARQ_URLSTOVISIT, "w") as f:
        while (not urls.empty()):
            link = urls.get()
            f.write(str(link) +"\n")

def loadArrayFromFile(arq):
    resp = []
    if (os.path.exists(arq)):
        with open(arq, "r") as f:
            for line in f:
                resp.append(line.strip())
    return resp

def loadQueueToVisit():
    return loadArrayFromFile(consts.ARQ_URLSTOVISIT)

def initQueueToVisit(queueUrlsVisitar):
    links = loadQueueToVisit()
    if len(links) > 0:
        for l in links:
            queueUrlsVisitar.put(l)
    else:
        fillQueueWithSeeds(queueUrlsVisitar)

def initkeywordsList(keywordsList: list):
    keywords = loadFileLikeArray(consts.ARQ_KEYWORDS)
    for k in keywords:
        keywordsList.append(k)

def initQueueUrlsVisited(queueUrlsVisited: list):
    urls = loadUrlsVisited()
    for u in urls:
        queueUrlsVisited.append(u)

def extractMetaFromPage(page):
    metas = {}
    metas["title"] = ""
    metas["keywords"] = ""
    metas["author"] = ""
    metas["lang"] = ""
    metas["description"] = ""

    soup = BeautifulSoup(page, "html.parser")
    if (soup.title is not None):
        metas["title"] = soup.title.string
    metas["lang"] = soup.find("html", "lang")
    meta = soup.find_all('meta')
    for tag in meta:
        if 'name' in tag.attrs.keys() and tag.attrs['name'].strip().lower() in ['description']:
            metas["description"] = tag.attrs['content']
        elif 'name' in tag.attrs.keys() and tag.attrs['name'].strip().lower() in ['keywords']:
            metas["keywords"] = tag.attrs['content']
        elif 'name' in tag.attrs.keys() and tag.attrs['name'].strip().lower() in ['author']:
            metas["author"] = tag.attrs['content']
        elif ('name' in tag.attrs.keys() and tag.attrs['name'].strip().lower() in ['lang']) or ('name' in tag.attrs.keys() and tag.attrs['name'].strip().lower() in ['language']):
            metas["lang"] = tag.attrs['content']
    return metas

def allUrlsFromDocument(data):
    """ Extrai urls da pagina

    Args:
        data (str): Pagina web

    Returns:
        list: urls encontradas
    """    
    extractor = URLExtract()
    urls = []
    try:
        urls = extractor.find_urls(data)
        return urls
    except:
        return urls

def allUrlsFromPage(page):
    """ Deprecated
        Extrai urls da pagina 

    Args:
        page (str): Pagina web

    Returns:
        list: urls encontradas
    """    
    soup = BeautifulSoup(page, "html.parser")
    links = []
    for link in soup.find_all('a'):
        links.append(link.get('href'))
    return links

def urlInBlackList(url):
    for domain in blackListDomains:
        if (domain in url):
            return True
    return False

def exportDataBaseToXlsx():
    wb = Workbook() 
    ws0 = wb.active 
    ws0.title = 'Craw-Homeopatia' 

    conn = sqlite3.connect(consts.ARQ_DATABASE) 
    cursor = conn.cursor() 
    cursor.execute("""  select *
                        from crawlerByDomain cbd  
                    """)
    result = cursor.fetchall()
    conn.commit()
    conn.close() 
    ws0.append(["id", "domain", "url", "page", "keys", "quantkeys", "titleFromPage", "keysFromPage", "langFromPage", "descFromPage", "authorFromPage"])
    for line in result: 
        ws0.append([line[0], line[1], line[2], line[3], line[4], line[5], line[6], line[7], line[8], line[9], line[10]])
    wb.save(consts.ARQ_XLSRESULT)

def countKeyWordsFrmDB():
    wb = Workbook() 
    ws0 = wb.active 
    ws0.title = 'Quant-Keywords'
    ws1 = wb.create_sheet(title='Quant-Key-FromPage')
    conn = sqlite3.connect(consts.ARQ_DATABASE) 
    cursor = conn.cursor() 
    cursor.execute("""  select keys, keysFromPage 
                        from crawlerByDomain cbd  
                    """)
    result = cursor.fetchall()
    conn.commit()
    conn.close() 
    colection = {}
    colectionFromPage = {}
    for keys in result:
        if (keys[0]):
            terms = keys[0].split(",")
            for term in terms:
                termNoSpace = term.strip()
                if termNoSpace in colection:
                    value = colection.get(termNoSpace)
                    colection[termNoSpace] = value + 1
                else:
                    colection[termNoSpace] = 1
            terms.clear()
        if (keys[1]):
            terms = keys[1].split(",")
            for term in terms:
                termNoSpace = term.strip()
                if termNoSpace in colectionFromPage: 
                    value = colectionFromPage.get(termNoSpace) 
                    colectionFromPage[termNoSpace] = value + 1 
                else:
                    colectionFromPage[termNoSpace] = 1 
            terms.clear()

    for chave, valor in colection.items():
        ws0.append([chave, valor]) 

    for chave, valor in colectionFromPage.items():
        ws1.append([chave, valor]) 
    
    saveTxtToCloud(colection)

    wb.save(consts.ARQ_XLSQUANTKEYS)

def saveTxtToCloud(colect):
    str = ""
    for chave, valor in colect.items():
        str = str + ' '.join([chave for i in range(valor)]) + ' '
    str = str.strip()
    with open(consts.ARQ_TXTFORCLOUD, 'w', encoding="utf-8") as text_file:
        text_file.write(str)

def madeCloudOfWords():
    d = os.path.dirname(__file__) if "__file__" in locals() else os.getcwd()
    text = open(os.path.join(d, consts.ARQ_TXTFORCLOUD)).read()

    # Generate a word cloud image
    wordcloud = WordCloud(width=1600, height=800, background_color="white", repeat=False, collocations=False)
    wordcloud.generate(text)
    wordcloud.to_file("keywordsFromVocabulary.png")
